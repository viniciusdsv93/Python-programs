# Program that creates a spreadsheet with a sum formula in the correct cell.

import openpyxl

# create a spreadsheet.

wb = openpyxl.Workbook()
sheet = wb['Sheet']

# save the sum formula in the correct cell.

sheet.cell(row = 3, column = 1).value = '=SUM(A1:A2)'

# ask the user to enter two numbers to be summed in the spreadsheet.

number1 = int(input('\nEnter the first number: '))
number2 = int(input('\nEnter the second number: '))

# save the user's answers in the correct cells to sum them using the formula.

sheet.cell(row = 1, column = 1).value = number1
sheet.cell(row = 2, column = 1).value = number2

# save the spreadsheet.

wb.save('sumformula.xlsx')
print('\nDone.\n')
