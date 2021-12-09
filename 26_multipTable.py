# Program that creates a spreadsheet as a multiplication table.

import openpyxl
from openpyxl.styles import Font

# create a spreadsheet.

wb = openpyxl.Workbook()

sheet = wb['Sheet']

# create the font to make it bold.

boldFont = Font(bold = True)

# ask the user for a number.

number = int(input('\nEnter a number: '))

# use a for loop to fill the spreadsheet cells with the number multiplicated.

for rowNum in range(1, number + 1):
    for columnNum in range(1, number + 1):
        sheet.cell(row = rowNum + 1, column = columnNum + 1).value = rowNum * columnNum

for rowNumHeader in range(1, number + 1):
    sheet.cell(row = rowNumHeader + 1, column = 1).font = boldFont
    sheet.cell(row = rowNumHeader + 1, column = 1).value = rowNumHeader

for columnNumHeader in range(1, number + 1):
    sheet.cell(row = 1, column = columnNumHeader + 1).font = boldFont
    sheet.cell(row = 1, column = columnNumHeader + 1).value = columnNumHeader


# save the spreadsheet as 'multTable.xlsx'.

wb.save('multTable.xlsx')
print('\nDone.\n')
