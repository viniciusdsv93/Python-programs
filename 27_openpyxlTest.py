import openpyxl

wb = openpyxl.load_workbook('example.xlsx')     # load a spreadsheet file. The file must exist before.

print(wb.sheetnames)        # print all the sheet names inside the file.
sheet = wb['Sheet1']        # call one of the sheets and store it in a variable.
print(sheet['A1'])          # print the type of the cell in row 1 column A.
print(sheet['A1'].value)    # print the value inside the cell.

c = sheet['B1']         # save a cell's value in a variable.
print(c.value)          # print the value inside the cell stored in the variable c.
print(f'''Row {c.row}    
column {c.column}
is {c.value}''')        # print the row and column where the c cell is located.

print(sheet['C1'].value)    # print the value stored in the cell named C1.

print(f'''Cell {c.coordinate}
is {c.value}.''')               # print the coordinate (with columns as letters).

print(sheet.cell(row = 1, column = 2).value)    # print the value stored in the specified cell.

for i in range(1, 8, 2):
    print(i, sheet.cell(row = i, column = 2).value)     # use a for loop to print only the values
                                                        # stored in the odd-numbered rows and column 2.

print(sheet.max_row)        # print the last row in the file's active sheet.
print(sheet.max_column)     # print the last column in the file's active sheet.

print(tuple(sheet['A1':'C3']))  # save multiple cells info in a tuple. 

for rowOfCellObjects in sheet['A1':'C3']:   # use a for loop to walk through the specified rows and
    for cellObj in rowOfCellObjects:        # columns, and print their coordinates and values.
        print(cellObj.coordinate, cellObj.value)
    print('End of row.')

sheet = wb.active

print(list(sheet.columns)[1])       # choose only the values inside the second column.

for cellObj in list(sheet.columns)[1]:      # use a for loop to print only the values on the second column.
    print(cellObj.value)

